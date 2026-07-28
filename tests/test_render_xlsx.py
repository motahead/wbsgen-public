import pytest
import io
from datetime import date
from datetime import datetime
import openpyxl
from wbsgen.planner import build_project_model
from wbsgen.render.xlsx import build_workbook
SAMPLE_DATA = {'project': {'name': 'XLSXテスト', 'startDate': '2026-07-13', 'endDate': '2026-07-24', 'statusDate': '2026-07-16'}, 'holidays': [{'date': '2026-07-17', 'name': '会社休日'}], 'tasks': [{'id': '1', 'name': '親タスク'}, {'id': '1.1', 'name': '子タスク', 'plannedStart': '2026-07-13', 'plannedDuration': 4, 'actualStart': '2026-07-14', 'actualEnd': None, 'progress': 50, 'issue': 65, 'comment': '実装中', 'assignee': '担当者A'}, {'id': '2', 'name': '単独タスク', 'plannedStart': '2026-07-20', 'plannedDuration': 3, 'progress': 0}]}

def build_result(data=None):
    return build_project_model(data or SAMPLE_DATA)

def roundtrip(workbook):
    """保存→読み戻しで、ファイルとして有効なことも同時に検証する。"""
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)

class TestWbsColumnsTests:

    def setup_method(self):
        result = build_result()
        assert not result.validation.has_errors
        self.book = roundtrip(build_workbook(result, source_label='sample.json'))
        self.sheet = self.book['WBS']

    def test_header_row_has_wbs_titles(self):
        titles = [self.sheet.cell(row=1, column=col).value for col in range(1, 15)]
        assert titles == ['ID', 'タスク名', '担当者', '計画開始', '計画終了', '実績開始', '実績終了', '進捗', '期待進捗', '差分', '遅れ(営業日)', '残り必要ペース', 'Issue', 'コメント']

    def test_task_rows_have_typed_values(self):
        assert self.sheet['A4'].value == '1.1'
        assert self.sheet['C4'].value == '担当者A'
        assert self.sheet['D4'].value == datetime(2026, 7, 13)
        assert self.sheet['D4'].number_format == 'yyyy/mm/dd'
        assert self.sheet['E4'].value == datetime(2026, 7, 16)
        assert self.sheet['F4'].value == datetime(2026, 7, 14)
        assert self.sheet['G4'].value is None
        assert self.sheet['H4'].value == 0.5
        assert self.sheet['H4'].number_format == '0%'
        assert self.sheet['I4'].value == 0.75
        assert self.sheet['J4'].value == -25
        assert self.sheet['K4'].value == 1
        assert self.sheet['L4'].value == 50.0
        assert self.sheet['M4'].value == 65
        assert self.sheet['N4'].value == '実装中'

    def test_assignee_is_blank_when_not_set(self):
        assert self.sheet['C3'].value is None

    def test_parent_row_is_bold_and_child_is_indented(self):
        assert self.sheet['B3'].font.bold
        assert not bool(self.sheet['B4'].font.bold)
        assert self.sheet['B3'].alignment.indent == 0
        assert self.sheet['B4'].alignment.indent == 1

    def test_wbs_detail_columns_have_nested_outline_levels(self):
        for column in ('A', 'B', 'P'):
            assert self.sheet.column_dimensions[column].outline_level == 0
            assert not self.sheet.column_dimensions[column].hidden
        for column in ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'M', 'N', 'O'):
            assert self.sheet.column_dimensions[column].outline_level == 1
            assert not self.sheet.column_dimensions[column].hidden
        for column in ('J', 'K', 'L'):
            assert self.sheet.column_dimensions[column].outline_level == 2
            assert not self.sheet.column_dimensions[column].hidden

    def test_freeze_panes_fixes_headers_and_wbs_columns(self):
        assert self.sheet.freeze_panes == 'P3'

    def test_comment_spacer_column_has_space_in_task_rows(self):
        for row in (3, 4, 5):
            assert self.sheet.cell(row=row, column=15).value == ' '

    def test_comment_spacer_column_is_blank_in_header_rows(self):
        assert self.sheet.cell(row=1, column=15).value is None
        assert self.sheet.cell(row=2, column=15).value is None

    def test_comment_spacer_column_width_and_outline(self):
        assert self.sheet.column_dimensions['O'].width == 2.0
        assert self.sheet.column_dimensions['O'].outline_level == 1
        assert not self.sheet.column_dimensions['O'].hidden

class TestDateHeaderTests:

    def setup_method(self):
        result = build_result()
        self.book = roundtrip(build_workbook(result))
        self.sheet = self.book['WBS']

    def test_month_header_is_merged_over_july(self):
        assert self.sheet['P1'].value == '2026年7月'
        merged = {str(range_) for range_ in self.sheet.merged_cells.ranges}
        assert 'P1:AA1' in merged

    def test_day_header_has_day_numbers(self):
        assert self.sheet['P2'].value == 13
        assert self.sheet['AA2'].value == 24

    def test_non_working_day_headers_are_shaded(self):
        assert self.sheet['T2'].fill.start_color.rgb == '00F2F5F8'
        assert self.sheet['U2'].fill.start_color.rgb == '00F2F5F8'
        assert self.sheet['R2'].fill.start_color.rgb == '00EEF2F6'

    def test_status_date_column_has_right_border_from_day_header_row(self):
        for row in (2, 3, 4, 5):
            border = self.sheet.cell(row=row, column=19).border
            assert border.right.style == 'thick'

    def test_date_columns_are_narrow(self):
        assert self.sheet.column_dimensions['P'].width == 4.0

class TestDateColumnHelperTests:

    def test_date_column_returns_first_subcolumn_by_default(self):
        from wbsgen.render.xlsx import DATE_FIRST_COL, _date_column
        dates = [date(2026, 7, 13), date(2026, 7, 14), date(2026, 7, 15)]
        assert _date_column(dates, date(2026, 7, 14), day_split=1) == DATE_FIRST_COL + 1

    def test_date_column_returns_right_edge_subcolumn(self):
        from wbsgen.render.xlsx import DATE_FIRST_COL, _date_column
        dates = [date(2026, 7, 13), date(2026, 7, 14), date(2026, 7, 15)]
        assert _date_column(dates, date(2026, 7, 14), day_split=2, right_edge=True) == DATE_FIRST_COL + 1 * 2 + 2 - 1

    def test_date_column_returns_none_when_out_of_range(self):
        from wbsgen.render.xlsx import _date_column
        dates = [date(2026, 7, 13), date(2026, 7, 14)]
        assert _date_column(dates, date(2026, 7, 20), day_split=1) is None

    def test_milestone_helpers_ignore_empty_and_out_of_range_dates(self):
        from wbsgen.render.xlsx import _filter_milestones_in_range
        from wbsgen.models import Milestone
        milestone = Milestone(date=date(2026, 7, 20), name='範囲外')
        assert _filter_milestones_in_range([milestone], []) == []
        assert _filter_milestones_in_range([milestone], [date(2026, 7, 13)]) == []

class TestDaySplitHeaderTests:

    def test_day_split_2_merges_each_day_and_halves_width(self):
        result = build_result()
        book = roundtrip(build_workbook(result, day_split=2))
        sheet = book['WBS']
        merged = {str(range_) for range_ in sheet.merged_cells.ranges}
        assert 'P2:Q2' in merged
        assert sheet['P2'].value == 13
        assert sheet.column_dimensions['P'].width == 2.0
        assert sheet.cell(row=2, column=23).border.right.style == 'thick'

    def test_invalid_day_split_is_rejected(self):
        with pytest.raises(ValueError):
            build_workbook(build_result(), day_split=3)

class TestOpenpyxlMissingTests:

    def test_build_workbook_raises_clear_error_without_openpyxl(self):
        from unittest import mock
        import wbsgen.render.xlsx as xlsx_module
        with mock.patch.object(xlsx_module, '_OPENPYXL_ERROR', ImportError("No module named 'openpyxl'")):
            with pytest.raises(ValueError) as context:
                build_workbook(build_result())
        assert 'openpyxl is required' in str(context.value)

class TestWorkbookEdgeCaseTests:

    def test_build_workbook_rejects_result_without_project_range(self):
        result = build_project_model({'project': None, 'tasks': []})
        with pytest.raises(ValueError, match='without a valid project'):
            build_workbook(result)

    def test_workbook_handles_unplanned_and_overdue_tasks(self):
        result = build_project_model({'project': {'name': 'P', 'startDate': '2026-06-01', 'endDate': '2026-06-10', 'statusDate': '2026-06-10'}, 'tasks': [{'id': '1', 'name': '未計画'}, {'id': '2', 'name': '遅延', 'plannedStart': '2026-06-01', 'plannedDuration': 1, 'progress': 0}]})
        book = roundtrip(build_workbook(result))
        sheet = book['WBS']
        assert sheet['I3'].value is None
        assert sheet['L4'].value == '達成不能'

    def test_milestone_helpers_ignore_columns_outside_the_date_range(self):
        from wbsgen.models import Milestone
        from wbsgen.render.xlsx import _apply_milestone_borders, _write_milestone_markers
        sheet = openpyxl.Workbook().active
        milestone = Milestone(date=date(2026, 6, 20), name='範囲外')
        dates = [date(2026, 6, 1)]
        _write_milestone_markers(sheet, [milestone], dates, 1, 3)
        _apply_milestone_borders(sheet, [milestone], dates, 1, 2, 3)
        assert sheet.cell(row=3, column=16).value is None

    def test_status_border_is_skipped_when_status_date_is_outside_range(self):
        result = build_project_model({'project': {'name': 'P', 'startDate': '2026-06-01', 'endDate': '2026-06-02', 'statusDate': '2026-06-03'}, 'tasks': []})
        assert build_workbook(result) is not None

class TestGanttCellTests:

    def setup_method(self):
        result = build_result()
        self.book = roundtrip(build_workbook(result))
        self.sheet = self.book['WBS']

    def rgb(self, coordinate):
        return self.sheet[coordinate].fill.start_color.rgb

    def test_child_progress_and_plan_fills(self):
        assert self.rgb('P4') == '004F936E'
        assert self.rgb('Q4') == '004F936E'
        assert self.rgb('R4') == '0092C8A6'
        assert self.rgb('S4') == '0092C8A6'

    def test_non_working_days_keep_holiday_fill_inside_plan(self):
        assert self.rgb('T4') == '00F2F5F8'
        assert self.rgb('U4') == '00F2F5F8'

    def test_zero_progress_task_has_plan_fill_only(self):
        for coordinate in ('W5', 'X5', 'Y5'):
            assert self.rgb(coordinate) == '0092C8A6'

    def test_actual_marks_cover_start_to_status_date(self):
        for coordinate in ('Q4', 'R4', 'S4'):
            assert self.sheet[coordinate].value == '■'
        assert self.sheet['P4'].value is None
        assert self.sheet['T4'].value is None
        assert self.sheet['Q4'].font.color.rgb == '002F3A4A'

    def test_parent_row_uses_parent_colors(self):
        assert self.rgb('P3') == '006689AD'
        assert self.rgb('R3') == '00A8BFD7'
        assert self.sheet['Q3'].font.color.rgb == '003F4A5A'

    def test_actual_mark_coexists_with_holiday_fill(self):
        data = {**SAMPLE_DATA, 'tasks': [dict(t) for t in SAMPLE_DATA['tasks']]}
        data['tasks'][1]['actualStart'] = '2026-07-16'
        data['tasks'][1]['actualEnd'] = '2026-07-18'
        book = roundtrip(build_workbook(build_result(data)))
        sheet = book['WBS']
        for coordinate in ('S4', 'T4', 'U4'):
            assert sheet[coordinate].value == '■'
        for coordinate in ('T4', 'U4'):
            assert sheet[coordinate].fill.start_color.rgb == '00F2F5F8'

class TestGanttDaySplitTests:

    def test_day_split_2_paints_half_day_boundary(self):
        data = {**SAMPLE_DATA, 'tasks': [dict(t) for t in SAMPLE_DATA['tasks']]}
        data['tasks'][1]['progress'] = 25
        book = roundtrip(build_workbook(build_result(data), day_split=2))
        sheet = book['WBS']
        assert sheet['P4'].fill.start_color.rgb == '004F936E'
        assert sheet['Q4'].fill.start_color.rgb == '004F936E'
        assert sheet['R4'].fill.start_color.rgb == '0092C8A6'
        for coordinate in ('R4', 'S4', 'T4', 'U4', 'V4', 'W4'):
            assert sheet[coordinate].value == '■'

class TestGanttClipTests:

    def test_plan_outside_display_range_is_clipped(self):
        data = {**SAMPLE_DATA, 'project': dict(SAMPLE_DATA['project'])}
        data['project']['startDate'] = '2026-07-14'
        result = build_result(data)
        book = roundtrip(build_workbook(result))
        sheet = book['WBS']
        assert sheet['P4'].value == '■'
        assert sheet['P4'].fill.start_color.rgb == '004F936E'

class TestInfoSheetTests:

    def test_info_sheet_has_project_metadata(self):
        book = roundtrip(build_workbook(build_result(), source_label='sample.json'))
        info = book['Info']
        assert info['A1'].value == 'プロジェクト名'
        assert info['B1'].value == 'XLSXテスト'
        assert info['A2'].value == '基準日'
        assert info['B2'].value == datetime(2026, 7, 16)
        assert info['A3'].value == '表示開始'
        assert info['B3'].value == datetime(2026, 7, 13)
        assert info['A4'].value == '表示終了'
        assert info['B4'].value == datetime(2026, 7, 24)
        assert info['A5'].value == '入力ファイル'
        assert info['B5'].value == 'sample.json'

    def test_info_sheet_shows_placeholder_without_source_label(self):
        book = roundtrip(build_workbook(build_result()))
        assert book['Info']['B5'].value == '-'

class TestMonthLabelSuppressionTests:

    def test_short_month_has_no_label_like_html(self):
        data = {**SAMPLE_DATA, 'project': dict(SAMPLE_DATA['project'])}
        data['project']['endDate'] = '2026-08-02'
        book = roundtrip(build_workbook(build_result(data)))
        sheet = book['WBS']
        assert sheet['P1'].value == '2026年7月'
        assert sheet.cell(row=1, column=35).value is None
        merged = {str(range_) for range_ in sheet.merged_cells.ranges}
        assert 'AI1:AJ1' in merged
        assert sheet.cell(row=1, column=35).fill.start_color.rgb == '00EEF2F6'

class TestActualMarkFontSizeTests:

    def test_actual_mark_font_size_shrinks_with_day_split(self):
        book1 = roundtrip(build_workbook(build_result()))
        assert book1['WBS']['Q4'].value == '■'
        assert book1['WBS']['Q4'].font.size == 11
        book2 = roundtrip(build_workbook(build_result(), day_split=2))
        assert book2['WBS']['R4'].value == '■'
        assert book2['WBS']['R4'].font.size == 8
        book4 = roundtrip(build_workbook(build_result(), day_split=4))
        assert book4['WBS']['T4'].value == '■'
        assert book4['WBS']['T4'].font.size == 5
MILESTONE_DATA = {**SAMPLE_DATA, 'milestones': [{'date': '2026-07-18', 'name': '中間レビュー'}]}

class TestMilestoneRowPlacementTests:

    def test_no_milestones_keeps_original_row_layout(self):
        book = roundtrip(build_workbook(build_result()))
        sheet = book['WBS']
        assert sheet['A4'].value == '1.1'
        assert sheet.freeze_panes == 'P3'

    def test_milestone_shifts_task_rows_and_freeze_panes(self):
        book = roundtrip(build_workbook(build_result(MILESTONE_DATA)))
        sheet = book['WBS']
        assert sheet['A5'].value == '1.1'
        assert sheet.freeze_panes == 'P4'

    def test_milestone_row_has_wbs_side_label(self):
        book = roundtrip(build_workbook(build_result(MILESTONE_DATA)))
        sheet = book['WBS']
        assert sheet['A3'].value == 'マイルストーン'
        merged = {str(range_) for range_ in sheet.merged_cells.ranges}
        assert 'A3:N3' in merged

    def test_milestone_row_shades_non_working_day_columns(self):
        book = roundtrip(build_workbook(build_result(MILESTONE_DATA)))
        sheet = book['WBS']
        assert sheet['T3'].fill.start_color.rgb == '00F2F5F8'
        assert sheet['U3'].fill.start_color.rgb == '00F2F5F8'
        assert sheet['R3'].fill.start_color.rgb == '00EEF2F6'

    def test_comment_spacer_column_excludes_milestone_row(self):
        book = roundtrip(build_workbook(build_result(MILESTONE_DATA)))
        sheet = book['WBS']
        assert sheet.cell(row=3, column=15).value is None
        for row in (4, 5, 6):
            assert sheet.cell(row=row, column=15).value == ' '

    def test_milestone_with_no_tasks_does_not_crash(self):
        data = {**SAMPLE_DATA, 'tasks': [], 'milestones': [{'date': '2026-07-18', 'name': '単独マイルストーン'}]}
        book = roundtrip(build_workbook(build_result(data)))
        sheet = book['WBS']
        assert sheet['A3'].value == 'マイルストーン'
        assert sheet.freeze_panes == 'P4'

class TestMilestoneMarkerTests:

    def test_single_milestone_marker_at_right_edge_of_day_column(self):
        book = roundtrip(build_workbook(build_result(MILESTONE_DATA)))
        sheet = book['WBS']
        assert sheet['U3'].value == '◆中間レビュー'
        assert sheet['U3'].font.color.rgb == '00C0392B'

    def test_milestone_vertical_line_spans_header_to_last_task_row(self):
        book = roundtrip(build_workbook(build_result(MILESTONE_DATA)))
        sheet = book['WBS']
        for row in (2, 3, 4, 5, 6):
            assert sheet.cell(row=row, column=21).border.right.style == 'thick'
            assert sheet.cell(row=row, column=21).border.right.color.rgb == '00C0392B'

    def test_same_date_milestones_are_joined_in_one_cell(self):
        data = {**SAMPLE_DATA, 'milestones': [{'date': '2026-07-18', 'name': '設計凍結'}, {'date': '2026-07-18', 'name': 'QA開始'}]}
        book = roundtrip(build_workbook(build_result(data)))
        sheet = book['WBS']
        assert sheet['U3'].value == '◆設計凍結、◆QA開始'

    def test_out_of_range_milestone_is_excluded(self):
        data = {**SAMPLE_DATA, 'milestones': [{'date': '2026-08-01', 'name': '範囲外'}]}
        book = roundtrip(build_workbook(build_result(data)))
        sheet = book['WBS']
        assert sheet['A4'].value == '1.1'
        assert sheet.freeze_panes == 'P3'

    def test_milestone_on_status_date_wins_border_color(self):
        data = {**SAMPLE_DATA, 'milestones': [{'date': '2026-07-16', 'name': '基準日と同日'}]}
        book = roundtrip(build_workbook(build_result(data)))
        sheet = book['WBS']
        assert sheet.cell(row=2, column=19).border.right.color.rgb == '00C0392B'

    def test_milestone_marker_on_non_working_day_keeps_holiday_fill(self):
        data = {**SAMPLE_DATA, 'milestones': [{'date': '2026-07-17', 'name': '会社休日と同日'}]}
        book = roundtrip(build_workbook(build_result(data)))
        sheet = book['WBS']
        assert sheet['T3'].value == '◆会社休日と同日'
        assert sheet['T3'].fill.start_color.rgb == '00F2F5F8'

    def test_milestone_marker_at_right_edge_with_day_split(self):
        book = roundtrip(build_workbook(build_result(MILESTONE_DATA), day_split=2))
        sheet = book['WBS']
        assert sheet['AA3'].value == '◆中間レビュー'

    def test_milestone_marker_with_no_tasks_does_not_crash(self):
        data = {**SAMPLE_DATA, 'tasks': [], 'milestones': [{'date': '2026-07-18', 'name': '単独マイルストーン'}]}
        book = roundtrip(build_workbook(build_result(data)))
        sheet = book['WBS']
        assert sheet['U3'].value == '◆単独マイルストーン'
