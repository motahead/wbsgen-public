"""Shared assertions for deterministic and exploratory zipapp workflows."""

from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import zipfile
from datetime import date
from pathlib import Path
from typing import Any


SOURCE_KEYS = ("project", "display", "tasks", "holidays", "milestones")
PANE_BOUNDARY_TOLERANCE_PX = 4
PANE_BOUNDARY_HIDDEN_STANDARD_COLUMNS = (
    "planned-period",
    "actual-period",
    "expected-progress",
    "comment",
)


def assert_pane_boundary_states(states: dict[str, dict[str, Any]]) -> None:
    """Ensure the left-pane edge and the resize divider remain contiguous."""
    for state_name, state in states.items():
        left_pane_right = (
            state.get("lastColumnRight", state["leftPaneRight"])
            if state_name == "analysis-after-standard-resize"
            else state["leftPaneRight"]
        )
        divider_x = state["dividerX"]
        gap = abs(left_pane_right - divider_x)
        if gap > PANE_BOUNDARY_TOLERANCE_PX:
            raise AssertionError(
                f"{state_name}: pane boundary gap={gap:g}px "
                f"(lastColumnRight={left_pane_right:g}, dividerX={divider_x:g})"
            )


def _pane_boundary_snapshot(page: Any) -> dict[str, Any]:
    return page.evaluate(
        """() => {
          const pane = document.querySelector('.left-pane');
          const handle = document.querySelector('.pane-resize-handle');
          const paneRect = pane.getBoundingClientRect();
          const handleRect = handle.getBoundingClientRect();
          const visibleHeads = Array.from(document.querySelectorAll('.left-head > .head-cell'))
            .filter((cell) => getComputedStyle(cell).display !== 'none');
          const lastColumnRight = visibleHeads.at(-1).getBoundingClientRect().right;
          const dividerOffset = Number.parseFloat(
            getComputedStyle(handle, '::before').left
          );
          return {
            leftPaneRight: paneRect.right,
            lastColumnRight,
            dividerX: handleRect.left + dividerOffset,
            activeView: document.documentElement.dataset.wbsView,
            standardColumns: Object.fromEntries(
              ['planned-period', 'actual-period', 'expected-progress', 'comment'].map(
                (column) => [column, document.querySelector(
                  `[data-column-visibility-toggle="${column}"]`
                ).checked]
              )
            ),
          };
        }"""
    )


def _drag_pane_handle(page: Any, delta_x: float) -> None:
    box = page.locator(".pane-resize-handle").bounding_box()
    if box is None:
        raise AssertionError("pane resize handle is not visible")
    start_x = box["x"] + box["width"] / 2
    start_y = box["y"] + min(box["height"] / 2, 24)
    page.mouse.move(start_x, start_y)
    page.mouse.down()
    page.mouse.move(start_x + delta_x, start_y, steps=8)
    page.mouse.up()


def capture_pane_boundary_states(html_path: Path) -> dict[str, dict[str, Any]]:
    """Capture the standard/analysis tab transitions relevant to Issue #167."""
    from playwright.sync_api import sync_playwright

    os.environ.setdefault(
        "PLAYWRIGHT_BROWSERS_PATH",
        str(Path(__file__).resolve().parent.parent / ".cache" / "ms-playwright"),
    )
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=True,
            args=["--disable-gpu", "--disable-dev-shm-usage"],
        )
        try:
            page = browser.new_page(viewport={"width": 1360, "height": 900})
            page.set_default_timeout(5000)
            page.goto(html_path.resolve().as_uri())
            page.wait_for_selector(".pane-resize-handle", timeout=5000)
            states = {"standard-initial": _pane_boundary_snapshot(page)}
            for column in PANE_BOUNDARY_HIDDEN_STANDARD_COLUMNS:
                page.locator(f'[data-column-visibility-toggle="{column}"]').evaluate(
                    "input => { if (input.checked) input.click(); }"
                )
            _drag_pane_handle(page, -80)
            states["standard-after-columns-and-resize"] = _pane_boundary_snapshot(page)
            page.locator('[data-wbs-view-target="analysis"]').click()
            page.wait_for_function("document.documentElement.dataset.wbsView === 'analysis'")
            states["analysis-after-standard-resize"] = _pane_boundary_snapshot(page)
            _drag_pane_handle(page, -80)
            states["analysis-after-resize"] = _pane_boundary_snapshot(page)
            page.locator('[data-wbs-view-target="standard"]').click()
            page.wait_for_function("document.documentElement.dataset.wbsView === 'standard'")
            states["standard-after-analysis-resize"] = _pane_boundary_snapshot(page)
            return states
        finally:
            browser.close()


def run_zipapp(zipapp: Path, args: list[str], cwd: Path) -> subprocess.CompletedProcess[str]:
    """Run a zipapp command and turn its diagnostics into an assertion failure."""
    completed = subprocess.run(
        [sys.executable, str(zipapp), *args],
        cwd=cwd,
        text=True,
        capture_output=True,
        check=False,
    )
    if completed.returncode:
        diagnostics = "\n".join(
            part
            for part in (
                f"zipapp command failed ({completed.returncode}): {' '.join(args)}",
                completed.stdout.strip(),
                completed.stderr.strip(),
            )
            if part
        )
        raise AssertionError(diagnostics)
    return completed


def _semantic_source(data: dict[str, Any]) -> dict[str, Any]:
    return {key: data[key] for key in SOURCE_KEYS if key in data}


def assert_source_equal(expected: dict[str, Any], actual: dict[str, Any]) -> None:
    """Assert equality of user-authored source fields, ignoring generator metadata."""
    expected_source = _semantic_source(expected)
    actual_source = _semantic_source(actual)
    differences = [
        key
        for key in SOURCE_KEYS
        if expected_source.get(key) != actual_source.get(key)
    ]
    if differences:
        raise AssertionError(
            "source fields differ: "
            + ", ".join(differences)
            + "\nexpected="
            + json.dumps(expected_source, ensure_ascii=False, sort_keys=True)
            + "\nactual="
            + json.dumps(actual_source, ensure_ascii=False, sort_keys=True)
        )


def is_weekend(day: date) -> bool:
    return day.weekday() >= 5


def leaf_tasks(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [task for task in data["tasks"] if "assignee" in task]


def derive_expected_warning_ids(data: dict[str, Any]) -> set[str]:
    holiday_dates = {holiday["date"] for holiday in data.get("holidays", [])}
    expected: set[str] = set()
    for task in leaf_tasks(data):
        if "plannedStart" not in task:
            expected.add(task["id"])
            continue
        planned_start = date.fromisoformat(task["plannedStart"])
        if is_weekend(planned_start) or task["plannedStart"] in holiday_dates:
            expected.add(task["id"])
    return expected


def derive_expected_leaf_plan_bar_ids(data: dict[str, Any]) -> set[str]:
    return {
        task["id"]
        for task in leaf_tasks(data)
        if "plannedStart" in task and "plannedDuration" in task
    }


def derive_expected_leaf_unplanned_ids(data: dict[str, Any]) -> set[str]:
    return {task["id"] for task in leaf_tasks(data) if "plannedStart" not in task}


def compare_dom_expectations(data: dict[str, Any], dom_state: dict[str, Any]) -> list[str]:
    differences: list[str] = []

    expected_warning_ids = derive_expected_warning_ids(data)
    actual_warning_ids = set(dom_state["warningRowIds"])
    if expected_warning_ids != actual_warning_ids:
        differences.append(
            "warning-row ids differ: "
            f"expected={sorted(expected_warning_ids)} actual={sorted(actual_warning_ids)}"
        )

    expected_plan_bar_ids = derive_expected_leaf_plan_bar_ids(data)
    expected_unplanned_ids = derive_expected_leaf_unplanned_ids(data)
    actual_plan_bar_ids = set(dom_state["planBarIds"])
    missing_plan_bars = expected_plan_bar_ids - actual_plan_bar_ids
    if missing_plan_bars:
        differences.append(f"missing plan bars for planned tasks: {sorted(missing_plan_bars)}")
    unexpected_plan_bars = expected_unplanned_ids & actual_plan_bar_ids
    if unexpected_plan_bars:
        differences.append(
            f"unplanned tasks unexpectedly have plan bars: {sorted(unexpected_plan_bars)}"
        )

    expected_milestone_count = len(data.get("milestones", []))
    if dom_state["milestoneCount"] != expected_milestone_count:
        differences.append(
            "milestone marker count differs: "
            f"expected={expected_milestone_count} actual={dom_state['milestoneCount']}"
        )

    expected_row_count = len(data["tasks"])
    if dom_state["rowCount"] != expected_row_count:
        differences.append(
            f"task row count differs: expected={expected_row_count} actual={dom_state['rowCount']}"
        )
    return differences


def assert_dom_matches_source(data: dict[str, Any], state: dict[str, Any]) -> None:
    differences = compare_dom_expectations(data, state)
    if differences:
        raise AssertionError("DOM expectations differ:\n- " + "\n- ".join(differences))


EXTRACTION_SCRIPT = """() => {
  const rows = Array.from(document.querySelectorAll('.left-rows .wbs-row'));
  const warningRowIds = rows
    .filter((row) => row.classList.contains('warning-row'))
    .map((row) => row.getAttribute('data-task-id'));
  const planBarIds = Array.from(
    document.querySelectorAll('[data-tooltip-role="plan-bar"]')
  ).map((bar) => bar.getAttribute('data-task-id'));
  const milestoneCount = document.querySelectorAll('.milestone-marker').length;
  return { rowCount: rows.length, warningRowIds, planBarIds, milestoneCount };
}"""


def capture_dom_state(html_path: Path) -> dict[str, Any]:
    from playwright.sync_api import sync_playwright

    os.environ.setdefault(
        "PLAYWRIGHT_BROWSERS_PATH",
        str(Path(__file__).resolve().parent.parent / ".cache" / "ms-playwright"),
    )
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=True,
            args=["--disable-gpu", "--disable-dev-shm-usage"],
        )
        try:
            page = browser.new_page(viewport={"width": 1360, "height": 900})
            page.goto(html_path.resolve().as_uri())
            page.wait_for_selector(".left-rows .wbs-row", timeout=5000)
            return page.evaluate(EXTRACTION_SCRIPT)
        finally:
            browser.close()


def assert_valid_xlsx(
    xlsx_path: Path,
    *,
    project_name: str | None = None,
    expected_wbs_rows: int | None = None,
) -> None:
    """Assert that an XLSX is structurally readable and has expected WBS essentials."""
    try:
        with zipfile.ZipFile(xlsx_path) as archive:
            corrupt_member = archive.testzip()
            names = set(archive.namelist())
    except (OSError, zipfile.BadZipFile) as exc:
        raise AssertionError(f"not a valid XLSX ZIP: {xlsx_path}: {exc}") from exc
    if corrupt_member is not None:
        raise AssertionError(f"corrupt XLSX member: {corrupt_member}")
    required = {"[Content_Types].xml", "xl/workbook.xml"}
    missing = required - names
    if missing:
        raise AssertionError(f"XLSX is missing required members: {sorted(missing)}")

    if project_name is None and expected_wbs_rows is None:
        return
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        raise AssertionError(f"cannot read XLSX workbook: {exc}") from exc
    try:
        if "WBS" not in workbook.sheetnames or "Info" not in workbook.sheetnames:
            raise AssertionError(f"unexpected XLSX sheets: {workbook.sheetnames}")
        if project_name is not None and workbook["Info"]["B1"].value != project_name:
            raise AssertionError(
                f"XLSX project name differs: expected={project_name!r} "
                f"actual={workbook['Info']['B1'].value!r}"
            )
        if expected_wbs_rows is not None:
            ids = [
                row[0]
                for row in workbook["WBS"].iter_rows(min_row=3, max_col=1, values_only=True)
                if isinstance(row[0], str) and re.fullmatch(r"\d+(?:\.\d+)*", row[0])
            ]
            if len(ids) != expected_wbs_rows:
                raise AssertionError(
                    "XLSX WBS row count differs: "
                    f"expected={expected_wbs_rows} actual={len(ids)}"
                )
    finally:
        workbook.close()
