"""XLSX export rendering for WBS-GEN."""

from __future__ import annotations

from datetime import date

from ..models import (
    MONTH_LABEL_MIN_DAYS,
    BuildResult,
    DisplayRow,
    Milestone,
    Project,
    WorkCalendar,
)
from ..planner import (
    expected_progress_for_task,
    flatten_computed_tasks,
    iter_dates,
    progress_analysis_for_task,
    working_dates_between,
)
from .tabular import PACE_UNATTAINABLE_LABEL, WBS_HEADERS

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string, get_column_letter

    _OPENPYXL_ERROR: ImportError | None = None
except ImportError as exc:  # pragma: no cover - 動作はユニットテストでmockして検証
    _OPENPYXL_ERROR = exc

__all__ = ["build_workbook"]

WBS_COLUMN_WIDTHS = (8, 30, 12, 11, 11, 11, 11, 8, 9, 8, 11, 13, 8, 30)
DETAIL_FIRST_COL_LETTER = "C"
DETAIL_LAST_COL_LETTER = "O"
ANALYSIS_FIRST_COL_LETTER = "J"
ANALYSIS_LAST_COL_LETTER = "L"
COMMENT_SPACER_COL = 15
COMMENT_SPACER_WIDTH = 2.0
DATE_FIRST_COL = 16
FIRST_TASK_ROW = 3
DAY_TOTAL_WIDTH = 4.0

COLOR_PLAN = "92C8A6"
COLOR_PROGRESS = "4F936E"
COLOR_ACTUAL_TEXT = "2F3A4A"
COLOR_PARENT_PLAN = "A8BFD7"
COLOR_PARENT_PROGRESS = "6689AD"
COLOR_PARENT_ACTUAL_TEXT = "3F4A5A"
COLOR_NON_WORKING = "F2F5F8"
COLOR_HEAD = "EEF2F6"
COLOR_STATUS_DATE = "DC2626"
COLOR_MILESTONE = "C0392B"

DATE_FORMAT = "yyyy/mm/dd"
PERCENT_FORMAT = "0%"
DELTA_FORMAT = '+0"pt";-0"pt";0"pt"'
DELAY_FORMAT = '0"日"'
PACE_FORMAT = '0.0"%/日"'
ACTUAL_MARK = "■"
# ■ が狭いサブ列からはみ出して基準日罫線を隠さないよう、分割数に応じて縮小する。
ACTUAL_MARK_FONT_SIZES = {1: 11, 2: 8, 4: 5}

if _OPENPYXL_ERROR is None:  # pragma: no branch - import guard is exercised by mock
    HEAD_FILL = PatternFill(
        start_color=COLOR_HEAD, end_color=COLOR_HEAD, fill_type="solid"
    )
    NON_WORKING_FILL = PatternFill(
        start_color=COLOR_NON_WORKING, end_color=COLOR_NON_WORKING, fill_type="solid"
    )
    PLAN_FILL = PatternFill(
        start_color=COLOR_PLAN, end_color=COLOR_PLAN, fill_type="solid"
    )
    PROGRESS_FILL = PatternFill(
        start_color=COLOR_PROGRESS, end_color=COLOR_PROGRESS, fill_type="solid"
    )
    PARENT_PLAN_FILL = PatternFill(
        start_color=COLOR_PARENT_PLAN, end_color=COLOR_PARENT_PLAN, fill_type="solid"
    )
    PARENT_PROGRESS_FILL = PatternFill(
        start_color=COLOR_PARENT_PROGRESS,
        end_color=COLOR_PARENT_PROGRESS,
        fill_type="solid",
    )
    BOLD_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")
    STATUS_DATE_BORDER = Border(right=Side(style="thick", color=COLOR_STATUS_DATE))
    MILESTONE_BORDER = Border(right=Side(style="thick", color=COLOR_MILESTONE))
    MILESTONE_FONT = Font(color=COLOR_MILESTONE, bold=True)


def _write_info_sheet(workbook, result: BuildResult, source_label: str | None) -> None:
    sheet = workbook.create_sheet("Info")
    project = result.project
    entries = (
        ("プロジェクト名", project.name, None),
        ("基準日", project.status_date, DATE_FORMAT),
        ("表示開始", result.display_start_date, DATE_FORMAT),
        ("表示終了", result.display_end_date, DATE_FORMAT),
        ("入力ファイル", source_label or "-", None),
    )
    for row_index, (label, value, number_format) in enumerate(entries, start=1):
        sheet.cell(row=row_index, column=1, value=label).font = BOLD_FONT
        cell = sheet.cell(row=row_index, column=2, value=value)
        if number_format is not None:
            cell.number_format = number_format
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 28


def build_workbook(
    result: BuildResult,
    *,
    source_label: str | None = None,
    day_split: int = 1,
):
    """Build an openpyxl Workbook from a computed project model."""
    if _OPENPYXL_ERROR is not None:
        raise ValueError(
            "openpyxl is required for XLSX export. Install dependencies with "
            "`python3 -m pip install -r requirements.txt` or use dist/wbsgen.pyz."
        )
    if day_split not in (1, 2, 4):
        raise ValueError("day_split must be 1, 2, or 4")
    project = result.project
    if (
        project is None
        or result.display_start_date is None
        or result.display_end_date is None
    ):
        raise ValueError("cannot export XLSX without a valid project and display range")

    calendar = WorkCalendar(holidays=tuple(result.holidays))
    dates = iter_dates(result.display_start_date, result.display_end_date)
    rows = flatten_computed_tasks(result.computed_roots)
    milestones = _filter_milestones_in_range(result.milestones, dates)
    first_task_row = FIRST_TASK_ROW + (1 if milestones else 0)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "WBS"
    _write_wbs_header(sheet)
    _write_date_header(sheet, dates, day_split, calendar)
    if milestones:
        _write_milestone_row(sheet, FIRST_TASK_ROW, dates, day_split, calendar)
        _write_milestone_markers(sheet, milestones, dates, day_split, FIRST_TASK_ROW)
    for index, display_row in enumerate(rows):
        row_index = first_task_row + index
        _write_wbs_cells(sheet, row_index, display_row, project, calendar)
        _paint_date_cells(
            sheet, row_index, display_row.task, project, calendar, dates, day_split
        )
    _write_comment_spacer_column(sheet, first_task_row, first_task_row + len(rows) - 1)
    _set_wbs_column_widths(sheet)
    _set_date_column_widths(sheet, len(dates), day_split)
    for index in range(
        column_index_from_string(DETAIL_FIRST_COL_LETTER),
        column_index_from_string(DETAIL_LAST_COL_LETTER) + 1,
    ):
        sheet.column_dimensions[get_column_letter(index)].outline_level = 1
    for index in range(
        column_index_from_string(ANALYSIS_FIRST_COL_LETTER),
        column_index_from_string(ANALYSIS_LAST_COL_LETTER) + 1,
    ):
        sheet.column_dimensions[get_column_letter(index)].outline_level = 2
    _apply_status_date_border(
        sheet, project, dates, day_split, first_task_row + len(rows) - 1
    )
    if milestones:
        _apply_milestone_borders(
            sheet, milestones, dates, day_split, 2, first_task_row + len(rows) - 1
        )
    sheet.freeze_panes = f"{get_column_letter(DATE_FIRST_COL)}{first_task_row}"
    _write_info_sheet(workbook, result, source_label)
    return workbook


def _date_column(
    dates: list[date],
    target: date,
    day_split: int,
    *,
    right_edge: bool = False,
) -> int | None:
    try:
        offset = dates.index(target)
    except ValueError:
        return None
    base_column = DATE_FIRST_COL + offset * day_split
    if right_edge:
        return base_column + day_split - 1
    return base_column


def _filter_milestones_in_range(
    milestones: list[Milestone], dates: list[date]
) -> list[Milestone]:
    if not dates:
        return []
    start, end = dates[0], dates[-1]
    in_range = [m for m in milestones if start <= m.date <= end]
    return [
        milestone
        for _, milestone in sorted(
            enumerate(in_range), key=lambda item: (item[1].date, item[0])
        )
    ]


def _write_milestone_row(
    sheet,
    row_index: int,
    dates: list[date],
    day_split: int,
    calendar: WorkCalendar,
) -> None:
    cell = sheet.cell(row=row_index, column=1, value="マイルストーン")
    cell.font = BOLD_FONT
    sheet.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index,
        end_column=len(WBS_HEADERS),
    )
    for offset, current in enumerate(dates):
        base_column = DATE_FIRST_COL + offset * day_split
        fill = NON_WORKING_FILL if calendar.is_non_working_day(current) else HEAD_FILL
        for sub in range(day_split):
            sheet.cell(row=row_index, column=base_column + sub).fill = fill


def _write_milestone_markers(
    sheet,
    milestones: list[Milestone],
    dates: list[date],
    day_split: int,
    row_index: int,
) -> None:
    grouped: dict[date, list[str]] = {}
    for milestone in milestones:
        grouped.setdefault(milestone.date, []).append(milestone.name)
    for milestone_date, names in grouped.items():
        column = _date_column(dates, milestone_date, day_split, right_edge=True)
        if column is None:
            continue
        text = "、".join(f"◆{name}" for name in names)
        cell = sheet.cell(row=row_index, column=column, value=text)
        cell.font = MILESTONE_FONT


def _apply_milestone_borders(
    sheet,
    milestones: list[Milestone],
    dates: list[date],
    day_split: int,
    first_row: int,
    last_row: int,
) -> None:
    columns = {
        _date_column(dates, milestone.date, day_split, right_edge=True)
        for milestone in milestones
    }
    for column in columns:
        if column is None:
            continue
        for row_index in range(first_row, last_row + 1):
            sheet.cell(row=row_index, column=column).border = MILESTONE_BORDER


def _write_wbs_header(sheet) -> None:
    for column, title in enumerate(WBS_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = BOLD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        sheet.merge_cells(
            start_row=1, start_column=column, end_row=2, end_column=column
        )


def _paint_date_cells(
    sheet,
    row_index: int,
    task,
    project: Project,
    calendar: WorkCalendar,
    dates: list[date],
    day_split: int,
) -> None:
    is_parent = bool(task.children)
    plan_fill = PARENT_PLAN_FILL if is_parent else PLAN_FILL
    progress_fill = PARENT_PROGRESS_FILL if is_parent else PROGRESS_FILL
    actual_font = Font(
        color=COLOR_PARENT_ACTUAL_TEXT if is_parent else COLOR_ACTUAL_TEXT,
        size=ACTUAL_MARK_FONT_SIZES[day_split],
    )

    planned_working: list[date] = []
    if task.planned_start is not None and task.planned_end is not None:
        planned_working = working_dates_between(
            task.planned_start, task.planned_end, calendar
        )
    dark_subcells = round(len(planned_working) * day_split * task.progress / 100)
    working_index = {value: index for index, value in enumerate(planned_working)}

    actual_start = task.actual_start
    actual_end: date | None = None
    if actual_start is not None:
        actual_end = (
            task.actual_end if task.actual_end is not None else project.status_date
        )

    for offset, current in enumerate(dates):
        base_column = DATE_FIRST_COL + offset * day_split
        non_working = calendar.is_non_working_day(current)
        planned_index = working_index.get(current)
        in_actual = (
            actual_start is not None and actual_start <= current <= actual_end
        )
        for sub in range(day_split):
            cell = sheet.cell(row=row_index, column=base_column + sub)
            if non_working:
                cell.fill = NON_WORKING_FILL
            elif planned_index is not None:
                global_subcell = planned_index * day_split + sub
                cell.fill = (
                    progress_fill if global_subcell < dark_subcells else plan_fill
                )
            if in_actual:
                cell.value = ACTUAL_MARK
                cell.font = actual_font
                cell.alignment = CENTER


def _write_wbs_cells(
    sheet,
    row_index: int,
    display_row: DisplayRow,
    project: Project,
    calendar: WorkCalendar,
) -> None:
    task = display_row.task
    is_parent = bool(task.children)
    sheet.cell(row=row_index, column=1, value=task.id)
    name_cell = sheet.cell(row=row_index, column=2, value=task.name)
    name_cell.alignment = Alignment(indent=display_row.depth)
    if is_parent:
        name_cell.font = BOLD_FONT
    if task.assignee:
        sheet.cell(row=row_index, column=3, value=task.assignee)
    date_values = (
        (4, task.planned_start),
        (5, task.planned_end),
        (6, task.actual_start),
        (7, task.source_task.actual_end),
    )
    for column, value in date_values:
        if value is not None:
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.number_format = DATE_FORMAT
    progress_cell = sheet.cell(row=row_index, column=8, value=task.progress / 100)
    progress_cell.number_format = PERCENT_FORMAT
    expected = expected_progress_for_task(task, project, calendar)
    if expected is not None:
        expected_cell = sheet.cell(row=row_index, column=9, value=expected / 100)
        expected_cell.number_format = PERCENT_FORMAT
    analysis = progress_analysis_for_task(task, project, calendar)
    if analysis.delta is not None:
        delta_cell = sheet.cell(row=row_index, column=10, value=analysis.delta)
        delta_cell.number_format = DELTA_FORMAT
    if analysis.delay_business_days is not None:
        delay_cell = sheet.cell(
            row=row_index, column=11, value=analysis.delay_business_days
        )
        delay_cell.number_format = DELAY_FORMAT
    if analysis.pace_unattainable:
        sheet.cell(row=row_index, column=12, value=PACE_UNATTAINABLE_LABEL)
    elif analysis.required_pace is not None:
        pace_cell = sheet.cell(row=row_index, column=12, value=analysis.required_pace)
        pace_cell.number_format = PACE_FORMAT
    if task.issue is not None:
        sheet.cell(row=row_index, column=13, value=task.issue)
    if task.comment:
        sheet.cell(row=row_index, column=14, value=task.comment)


def _write_comment_spacer_column(
    sheet, first_task_row: int, last_task_row: int
) -> None:
    for row_index in range(first_task_row, last_task_row + 1):
        sheet.cell(row=row_index, column=COMMENT_SPACER_COL, value=" ")
    sheet.column_dimensions[
        get_column_letter(COMMENT_SPACER_COL)
    ].width = COMMENT_SPACER_WIDTH


def _set_wbs_column_widths(sheet) -> None:
    for index, width in enumerate(WBS_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_date_header(sheet, dates: list[date], day_split: int, calendar: WorkCalendar) -> None:
    month_first = 0
    for offset in range(1, len(dates) + 1):
        is_boundary = offset == len(dates) or dates[offset].month != dates[month_first].month
        if not is_boundary:
            continue
        first_column = DATE_FIRST_COL + month_first * day_split
        last_column = DATE_FIRST_COL + offset * day_split - 1
        label_date = dates[month_first]
        cell = sheet.cell(row=1, column=first_column)
        # HTMLと同様、表示範囲内の日数が MONTH_LABEL_MIN_DAYS 未満の月はラベルを出さない。
        if offset - month_first >= MONTH_LABEL_MIN_DAYS:
            cell.value = f"{label_date.year}年{label_date.month}月"
        cell.font = BOLD_FONT
        cell.fill = HEAD_FILL
        if last_column > first_column:
            sheet.merge_cells(
                start_row=1, start_column=first_column, end_row=1, end_column=last_column
            )
        month_first = offset
    for offset, current in enumerate(dates):
        base_column = DATE_FIRST_COL + offset * day_split
        cell = sheet.cell(row=2, column=base_column, value=current.day)
        cell.alignment = CENTER
        cell.fill = (
            NON_WORKING_FILL if calendar.is_non_working_day(current) else HEAD_FILL
        )
        if day_split > 1:
            sheet.merge_cells(
                start_row=2,
                start_column=base_column,
                end_row=2,
                end_column=base_column + day_split - 1,
            )


def _apply_status_date_border(
    sheet, project: Project, dates: list[date], day_split: int, last_row: int
) -> None:
    column = _date_column(dates, project.status_date, day_split, right_edge=True)
    if column is None:
        return
    # 月ヘッダー(1行目)は結合セルのため内部に縦罫線を描画できない。日ヘッダー行(2行目)以降に適用する。
    for row_index in range(2, last_row + 1):
        sheet.cell(row=row_index, column=column).border = STATUS_DATE_BORDER


def _set_date_column_widths(sheet, date_count: int, day_split: int) -> None:
    sub_width = DAY_TOTAL_WIDTH / day_split
    for offset in range(date_count * day_split):
        sheet.column_dimensions[
            get_column_letter(DATE_FIRST_COL + offset)
        ].width = sub_width
